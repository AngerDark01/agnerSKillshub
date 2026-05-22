"""
scripts/xlsx/table_tools.py
===========================
读取 Excel 附件中的表格数据，供 agent 自主判断列名和提取目标行。

设计原则：
    - 不用正则匹配列名，让 agent 看到完整列名后自己判断
    - 提取时 agent 指定列序号，避免跨文件列名不统一的脆性问题
    - 多行汇总时支持去重和合并，处理一个标包对应多条记录的情况

工具一：inspect_xlsx(path)
    列出所有 sheet 的列名（带序号）和前几行数据，供 agent 理解结构。

工具二：extract_xlsx_rows(path, sheet_name, match, header_row_idx)
    按 {列序号: 值} 条件提取匹配行，返回字段字典列表。
    agent 先 inspect 看列名和序号，再决定用哪列做匹配。

工具三：summarize_xlsx_field(path, sheet_name, match, target_cols, header_row_idx)
    从多匹配行中提取指定列的值，去重后汇总为可读字符串。
    适合"一个标包多个交货地点/多个交货日期"的情况。
"""

import openpyxl
from pathlib import Path


# ══════════════════════════════════════════════════════════════
# 工具一：inspect_xlsx
# ══════════════════════════════════════════════════════════════

def inspect_xlsx(path: str, preview_rows: int = 3) -> dict:
    """
    读取 Excel 文件结构：sheet 列表、列名（带序号）、前几行数据预览。

    返回：
        {
          "file": str,
          "sheets": [
            {
              "name":            str,
              "total_rows":      int,   # 含表头
              "total_cols":      int,
              "header_row_idx":  int,   # 表头所在行（0-based）
              "headers":         list[str],  # 原始列名
              "headers_indexed": list[str],  # "[0]列名" 格式，方便 agent 引用
              "preview":         list[dict]  # 前 N 行，字段字典
            }
          ]
        }
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)
    sheets_info = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        all_rows = [list(r) for r in ws.iter_rows(values_only=True)]

        # 去掉末尾完全空行
        while all_rows and not any(v is not None for v in all_rows[-1]):
            all_rows.pop()

        if not all_rows:
            continue

        # 找第一个有内容的行作为表头
        header_row_idx = 0
        for i, row in enumerate(all_rows):
            if any(v is not None for v in row):
                header_row_idx = i
                break

        raw_headers = [str(h) if h is not None else "" for h in all_rows[header_row_idx]]
        headers_indexed = [f"[{i}]{h}" for i, h in enumerate(raw_headers)]

        # 数据预览
        preview = []
        data_rows = all_rows[header_row_idx + 1:]
        for row in data_rows[:preview_rows]:
            if any(v is not None for v in row):
                d = {}
                for i, h in enumerate(raw_headers):
                    v = row[i] if i < len(row) else None
                    d[h or f"col_{i}"] = str(v) if v is not None else ""
                preview.append(d)

        sheets_info.append({
            "name":            sheet_name,
            "total_rows":      len(all_rows),
            "total_cols":      len(raw_headers),
            "header_row_idx":  header_row_idx,
            "headers":         raw_headers,
            "headers_indexed": headers_indexed,
            "preview":         preview,
        })

    wb.close()
    return {"file": str(path), "sheets": sheets_info}


# ══════════════════════════════════════════════════════════════
# 工具二：extract_xlsx_rows
# ══════════════════════════════════════════════════════════════

def extract_xlsx_rows(
    path: str,
    sheet_name: str,
    match: dict,
    header_row_idx: int = 0,
) -> dict:
    """
    按条件提取匹配行，返回字段字典列表。

    参数：
        path:           xlsx 文件路径
        sheet_name:     sheet 名称（从 inspect_xlsx 获取）
        match:          匹配条件，key 为列序号（int），value 为目标值（str）
                        支持多条件 AND，值比较忽略首尾空白
                        示例：{0: "282602-1104024-9999", 1: "包1"}
        header_row_idx: 表头所在行（0-based，从 inspect_xlsx 获取，默认 0）

    返回：
        {
          "status": "ok" | "not_found" | "ambiguous",
          "count":  int,
          "headers": list[str],
          "rows":    list[dict],   # 每行为 {列名: 值} 字典
          "detail":  str
        }
    """
    wb = openpyxl.load_workbook(str(path), data_only=True)

    if sheet_name not in wb.sheetnames:
        wb.close()
        return {"status": "not_found", "count": 0, "headers": [], "rows": [],
                "detail": f"sheet '{sheet_name}' 不存在，可用: {wb.sheetnames}"}

    ws = wb[sheet_name]
    all_rows = [list(r) for r in ws.iter_rows(values_only=True)]
    wb.close()

    if header_row_idx >= len(all_rows):
        return {"status": "not_found", "count": 0, "headers": [], "rows": [],
                "detail": f"header_row_idx={header_row_idx} 超出范围"}

    headers = [str(h) if h is not None else "" for h in all_rows[header_row_idx]]

    hits = []
    for row in all_rows[header_row_idx + 1:]:
        # 跳过完全空行
        if not any(v is not None for v in row):
            continue

        match_all = True
        for col_idx, val in match.items():
            cell = str(row[col_idx]).strip() if col_idx < len(row) and row[col_idx] is not None else ""
            if str(val).strip() != cell:
                match_all = False
                break

        if match_all:
            d = {}
            for i, h in enumerate(headers):
                v = row[i] if i < len(row) else None
                d[h or f"col_{i}"] = str(v) if v is not None else ""
            hits.append(d)

    if not hits:
        return {"status": "not_found", "count": 0, "headers": headers, "rows": [],
                "detail": f"未找到匹配行，条件: {match}"}

    status = "ok" if len(hits) >= 1 else "not_found"
    return {"status": status, "count": len(hits), "headers": headers,
            "rows": hits, "detail": f"找到 {len(hits)} 行"}


# ══════════════════════════════════════════════════════════════
# 工具三：summarize_xlsx_field
# ══════════════════════════════════════════════════════════════

def summarize_xlsx_field(
    path: str,
    sheet_name: str,
    match: dict,
    target_cols: list,
    header_row_idx: int = 0,
    dedup: bool = True,
) -> dict:
    """
    从多匹配行中提取指定列的值，去重后汇总。

    适用场景：一个标包/包对应多条记录（如多个交货地点、多个交货日期），
    需要汇总为一个可读字符串写入文档。

    参数：
        path, sheet_name, match, header_row_idx: 同 extract_xlsx_rows
        target_cols:  要提取的列序号列表，如 [11, 12, 13]
        dedup:        是否对每列的值去重（默认 True）

    返回：
        {
          "status": "ok" | "not_found",
          "matched_rows": int,
          "fields": {
            "列名": {
              "values":   list[str],   # 去重后的值列表
              "summary":  str          # 以"；"连接的汇总字符串
            }
          },
          "detail": str
        }
    """
    result = extract_xlsx_rows(path, sheet_name, match, header_row_idx)
    if result["status"] == "not_found":
        return {"status": "not_found", "matched_rows": 0, "fields": {},
                "detail": result["detail"]}

    headers = result["headers"]
    rows    = result["rows"]

    fields = {}
    for col_idx in target_cols:
        if col_idx >= len(headers):
            continue
        col_name = headers[col_idx]
        vals = [r.get(col_name, "") for r in rows]
        vals = [v for v in vals if v and v.lower() not in ("none", "")]
        if dedup:
            seen = []
            for v in vals:
                if v not in seen:
                    seen.append(v)
            vals = seen
        fields[col_name] = {
            "values":  vals,
            "summary": "；".join(vals) if vals else "（无数据）",
        }

    return {
        "status":       "ok",
        "matched_rows": result["count"],
        "fields":       fields,
        "detail":       f"共 {result['count']} 行，提取 {len(target_cols)} 个字段",
    }
